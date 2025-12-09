"""
Excel Generator - Generate Excel reports with formatting
"""
from typing import Any, Dict, List, Optional, Union
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from src.generators.base_generator import BaseGenerator
from src.core.config import settings


class ExcelGenerator(BaseGenerator):
    """
    Generator for Excel reports with professional formatting
    """

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize Excel generator

        Args:
            output_dir: Output directory
        """
        super().__init__(output_dir or settings.EXCEL_DIR)
        
        # Default styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_alignment = Alignment(horizontal='center', vertical='center')

    def generate(self, data: pd.DataFrame, filename: str, 
                template: str = None, **kwargs) -> Path:
        """
        Generate Excel report

        Args:
            data: DataFrame with report data
            filename: Output filename
            template: Optional template name
            **kwargs: Additional options (sheet_name, add_timestamp, etc.)

        Returns:
            Path to generated file
        """
        if not self.validate_data(data):
            raise ValueError("Invalid data provided")

        filename = self._prepare_filename(filename, '.xlsx')
        
        if kwargs.get('add_timestamp', False):
            filename = self._add_timestamp_to_filename(filename)

        output_path = self._get_output_path(filename, 'excel')
        sheet_name = kwargs.get('sheet_name', 'Report')

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply header styling
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.header_alignment
                
                cell.border = self.border

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Add charts if configured
        if 'charts' in kwargs:
            self._add_charts(ws, kwargs['charts'], len(data))

        wb.save(output_path)
        self.generated_files.append(output_path)
        self.logger.info(f"Excel report generated: {output_path}")
        
        return output_path

    def generate_multi_sheet(self, data_dict: Dict[str, pd.DataFrame], 
                            filename: str, **kwargs) -> Path:
        """
        Generate Excel with multiple sheets

        Args:
            data_dict: Dictionary of sheet_name: DataFrame
            filename: Output filename
            **kwargs: Additional options

        Returns:
            Path to generated file
        """
        filename = self._prepare_filename(filename, '.xlsx')
        output_path = self._get_output_path(filename, 'excel')

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for sheet_name, data in data_dict.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

            for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    if r_idx == 1:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.header_alignment
                    
                    cell.border = self.border

            self._auto_adjust_columns(ws)

        wb.save(output_path)
        self.generated_files.append(output_path)
        self.logger.info(f"Multi-sheet Excel report generated: {output_path}")
        
        return output_path

    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_charts(self, worksheet, chart_configs: List[Dict], data_rows: int) -> None:
        """
        Add charts to worksheet

        Args:
            worksheet: Target worksheet
            chart_configs: List of chart configurations
            data_rows: Number of data rows
        """
        for idx, config in enumerate(chart_configs):
            chart_type = config.get('type', 'bar')
            title = config.get('title', 'Chart')
            x_col = config.get('x', 1)
            y_col = config.get('y', 2)

            if chart_type == 'bar':
                chart = BarChart()
            elif chart_type == 'line':
                chart = LineChart()
            elif chart_type == 'pie':
                chart = PieChart()
            else:
                continue

            chart.title = title

            # Data reference
            data = Reference(worksheet, min_col=y_col, min_row=1, 
                           max_row=data_rows + 1, max_col=y_col)
            categories = Reference(worksheet, min_col=x_col, min_row=2, 
                                  max_row=data_rows + 1)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Position chart
            chart_position = f"A{data_rows + 5 + (idx * 15)}"
            worksheet.add_chart(chart, chart_position)

    def set_header_style(self, font: Font = None, fill: PatternFill = None) -> None:
        """
        Set custom header style

        Args:
            font: Font style
            fill: Fill style
        """
        if font:
            self.header_font = font
        if fill:
            self.header_fill = fill

    def add_summary_row(self, worksheet, data: pd.DataFrame, 
                       sum_columns: List[str]) -> None:
        """
        Add summary row to worksheet

        Args:
            worksheet: Target worksheet
            data: Source DataFrame
            sum_columns: Columns to sum
        """
        last_row = worksheet.max_row + 1
        
        worksheet.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
        
        for col_idx, col_name in enumerate(data.columns, 1):
            if col_name in sum_columns:
                total = data[col_name].sum()
                cell = worksheet.cell(row=last_row, column=col_idx, value=total)
                cell.font = Font(bold=True)
                cell.border = self.border
